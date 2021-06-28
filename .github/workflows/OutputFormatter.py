from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Border, Side
import win32com.client as win32
import os
import time


class OutputFormatter:

    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(filename=path)
        self.ws = self.wb.active
        self.header_row_idx = "8"

    def processor(self, fiscal_year):

        # Modify sheet name
        self.ws.title = 'POS AUDIT TIE OUT Q' + fiscal_year[1]

        # Write In Scope and Quarter
        self.ws["2"][2].value = 'In Scope?'
        self.ws["2"][3].value = fiscal_year

        # Border Styles
        medium = Side(border_style="medium", color="000000")

        # Color Fills
        yellowFill = PatternFill(start_color='fff2cc', end_color='fff2CC', fill_type='solid')
        cyanFill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        oceanBlueFill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        lightGreenFill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        # Font Styles
        thin_font = Font(name="Calibri", size=8)
        thick_font = Font(name="Calibri",  size=9, bold=True)
        total_row_font = Font(name="Calibri",  size=8, bold=True)

        # Format Header Row. Formatting has to be done for each cell.
        row = self.ws[self.header_row_idx]       # Get header row.
        for cell in row:
            cell.font = thick_font

        # List of columns that require number formatting.
        formatting_col_idx = [1, 3, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]

        # Format each data cell -
        for data_row_idx in range(9, self.ws.max_row+1):
            row = self.ws[str(data_row_idx)]  # Get data row.
            total_row = "Total" in row[0].value
            for cell_idx, cell in enumerate(row):

                if not total_row:
                    cell.font = thin_font

                    if cell_idx in formatting_col_idx:
                        cell.number_format = '0'

                    # Apply Borders
                    if cell_idx == 22 or cell_idx == 27 or cell_idx == 33:
                        cell.border = Border(left=medium)

                    # Apply Yellow Fill to Model N columns
                    if 22 <= cell_idx < 27:
                        cell.fill = yellowFill

                    # Apply Cyan Fill to Global POS Team Input columns
                    if 27 <= cell_idx < 33:
                        cell.fill = cyanFill
                else:
                    cell.font = total_row_font

                    if cell_idx == 22 or cell_idx == 27 or cell_idx == 33:
                        cell.border = Border(top=medium, bottom=medium, left=medium)
                    else:
                        cell.border = Border(top=medium, bottom=medium)

                    if 1 < cell_idx < self.ws.max_column-2:
                        cell.fill = oceanBlueFill

                    if cell_idx >= self.ws.max_column-2:
                        cell.fill = lightGreenFill

                    if cell_idx == self.ws.max_column-1:
                        cell.border = Border(top=medium, bottom=medium, right=medium)

    def data_grouper(self, df):

        xlApp = win32.Dispatch("Excel.Application")
        wb = xlApp.Workbooks.Open(os.path.join(os.getcwd(), self.path))
        ws = wb.Worksheets(1)
        xlApp.Visible = False

        total_rows_idx = [idx for idx, i in enumerate(df['LookUp']) if 'Grand Total' not in i and 'Total' in i]
        last_total_idx = 9

        for total_idx in total_rows_idx:
            actual_row_idx = total_idx + 9
            for row_idx in range(last_total_idx, actual_row_idx):
                ws.Rows(row_idx).OutlineLevel = 2
            ws.Rows(actual_row_idx).OutlineLevel = 1
            rangeObj = ws.Range(f"A{last_total_idx}:A{actual_row_idx}")
            last_total_idx = actual_row_idx + 1
            rangeObj.EntireRow.Group

        ws.Outline.ShowLevels(RowLevels=1)

        wb.RefreshAll()
        time.sleep(5)
        wb.Save()

        wb.Close(False)
        xlApp.Quit

    def file_saver(self):
        self.wb.save(self.path)


if __name__ == '__main__':
    file_path = "Output Files/Audit Process Report.xlsx"
    import pandas as pd
    data = pd.read_excel(file_path, skiprows=7)
    obj = OutputFormatter(file_path)
    obj.file_saver()
    obj.data_grouper(data)
